import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill

tabulationSheetOne = pd.read_csv('results_report (18).csv')
philoComments = pd.read_csv('raw_scores_report (17).csv')
sisterhoodComments = pd.read_csv('raw_scores_report (12).csv')
toursComments=pd.read_csv('raw_scores_report (19).csv')
prefComments = pd.read_csv('raw_scores_report (20).csv')
tabulationSheetOne.columns = tabulationSheetOne.columns.str.strip()
tabulationSheetOneFixed = tabulationSheetOne.drop('List', axis=1, errors='ignore')

keyColumns = ['First Name', 'Last Name']
commentColumnName = 'Comment'
interestColumnName = 'AOII Interest'

def processComments(df, prefix):
    if not all(col in df.columns for col in keyColumns + [commentColumnName] + [interestColumnName]):
        return pd.DataFrame(columns=keyColumns)
    
    subset = df[keyColumns + [commentColumnName] + [interestColumnName]].copy()
    for col in keyColumns:
        subset[col] = subset[col].astype(str).str.strip()
    
    subset['comment_num'] = subset.groupby(keyColumns).cumcount() + 1

    pivot = subset.pivot_table(index = keyColumns, columns = 'comment_num', values = [commentColumnName, interestColumnName], aggfunc = 'first')
    pivot.columns = [f'{prefix} {level0} {level1}' for level0, level1 in pivot.columns]

    return pivot.reset_index()

sisterhoodCommentPivot = processComments(sisterhoodComments, 'Sisterhood')
philoCommentPivot = processComments(philoComments, 'Philanthropy')
chapterCommentPivot = processComments(toursComments, 'Chapter Tours')
prefCommentPivot = processComments(prefComments, 'Pref')

columnsToKeep = ['Council ID', 'First Name', 'Last Name', 'Overall', 'AOII Interest (0)', 'Ambition (0)', 'Likability (0)', 'Sisterhood', 'Philanthropy', 'Chapter Tours', 'Preference','Pool']
columnsToRound = ['Overall', 'AOII Interest (0)', 'Ambition (0)', 'Likability (0)', 'Sisterhood', 'Philanthropy', 'Chapter Tours', 'Preference']

existingColumnsToKeep = [col for col in columnsToKeep if col in tabulationSheetOneFixed.columns]
newSheet = tabulationSheetOneFixed[existingColumnsToKeep].copy()

for col in columnsToRound:
    if col in newSheet.columns:
        newSheet[col] = newSheet[col].apply(lambda x: np.sign(x) * np.floor(np.abs(x) + 0.5) if pd.notna(x) else x)

for col in columnsToRound:
    if col in newSheet.columns:
        newSheet[col] = pd.to_numeric(newSheet[col], errors='coerce')

filteredSheet = newSheet[newSheet['Overall'].notna() & (newSheet['Overall'] != 0)].copy()
filteredSheet = filteredSheet.sort_values(by='Overall', ascending=False)

if 'Sisterhood' in filteredSheet.columns:
    sisterhoodRoundOne = filteredSheet[filteredSheet['Sisterhood'].notna()].copy()
else:
    sisterhoodRoundOne = pd.DataFrame()

mergedSheet = pd.merge( # merges filtered and commented sheet to get all filtered cols but add comments
    filteredSheet,
    sisterhoodCommentPivot,
    on = keyColumns,
    how = 'left'
)

mergedSheet = pd.merge(
    mergedSheet,
    philoCommentPivot,
    on=keyColumns,
    how='left'
)

mergedSheet = pd.merge(
    mergedSheet,
    chapterCommentPivot,
    on=keyColumns,
    how='left'
)

mergedSheet = pd.merge(
    mergedSheet,
    prefCommentPivot,
    on=keyColumns,
    how='left'
)

philanthropyPrimary = pd.DataFrame()
philantrhopySecondary = pd.DataFrame()

if 'Philanthropy' in mergedSheet.columns and 'Sisterhood' in mergedSheet.columns and 'Pool' in mergedSheet.columns:
    philanthropyAll = mergedSheet[(mergedSheet['Philanthropy'].notna()) & (mergedSheet['Sisterhood'].notna()) & (mergedSheet['Chapter Tours'].notna()) & (mergedSheet['Preference'].notna())].copy()
    philanthropyPrimary = philanthropyAll[philanthropyAll['Pool']== 'Primary'].copy()
    philanthropySecondary = philanthropyAll[philanthropyAll['Pool']== 'Secondary'].copy()

def randomNames(masterDataFrame):
    fake = Faker()

    masterNameList = pd.DataFrame()
    for df in masterDataFrame:
        if 'First Name' in df.columns and 'Last Name' in df.columns:
            masterNameList = pd.concat([masterNameList, df[['First Name', 'Last Name']]])

    uniquePeople = masterNameList.drop_duplicates().reset_index(drop=True)

    name_map = {}
    for index, row in uniquePeople.iterrows():
        real_key = (row['First Name'], row['Last Name'])


        fakeFirst = fake.first_name_female()
        fakeLast = fake.last_name()

        name_map[real_key] = (fakeFirst, fakeLast)

    for df in masterDataFrame:
        real_key = (row['First Name'], row['Last Name'])

        in 'First Name' in df.columns and 'Last Name' in df.columns:
        newFirstNames = []
        newLastNames = []

        for index, row if df.iterrows():
            if real_key in name_map:
                fake_f, fake_l = name_map[real_key]
                newFirstNames.append(fake_f)
                newLastNames.append(fake_l)
            else:
                newFirstNames.append("Unknown")
                newLastNames.append("Unknown")
        
        df['First Name'] = newFirstNames
        df['Last Name'] = newLastNames

    print(masterDataFrame[0][['First Name', 'Last Name']].head())

newDataFrame = randomNames(philanthropyAll)

with pd.ExcelWriter('Scores.xlsx', engine='openpyxl') as writer: # CHANGE BASED ON ROUND
    newSheet.to_excel(writer, sheet_name="All Girls MasterList", index=False)
    if not philanthropyPrimary.empty:
        philanthropyPrimary.to_excel(writer, sheet_name='Primary', index=False)
    if not philanthropySecondary.empty:
        philanthropySecondary.to_excel(writer, sheet_name='Secondary', index=False)

    workbook = writer.book
    greenFill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type="solid")
    lightGreenFill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type="solid")
    purpleFill = PatternFill(start_color='C9A0DC', end_color='C9A0DC', fill_type="solid")
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type="solid")
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")
    whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")

    for sheetName in writer.sheets:
        worksheet = writer.sheets[sheetName]

        headers = [cell.value for cell in worksheet[1]]
        overallCol, sisterhoodCol, philoCol = None, None, None

        try:
            overallCol = headers.index('Overall')
            if 'Philanthropy' in headers: philoCol = headers.index('Philanthropy')
        except ValueError:
            continue

        if overallCol is not None:
            for row in worksheet.iter_rows(min_row=2):
                overallCell = row[overallCol]
                score = 0

                if overallCell.value is not None:
                    try:
                        score = float(overallCell.value)
                    except (ValueError, TypeError):
                        continue

                if sheetName == 'Primary' or sheetName == 'Secondary': # philo sheet coloring
                    if philoCol is not None:
                        philoCell = row[philoCol]
                        isPhiloEmpty = philoCell.value is None or philoCell.value == ''

                        if isPhiloEmpty: 
                            for cell in row: cell.fill = redFill
                        else:
                            if score >= 22:
                                for cell in row: cell.fill = purpleFill
                            if 8 <= score < 22:
                                for cell in row: cell.fill = greenFill
                            elif 6 <= score < 8:
                                for cell in row: cell.fill = lightGreenFill
                            elif score < 6:
                                for cell in row: cell.fill = yellowFill

                if sheetName == 'All Girls MasterList': # masterlist coloring
                    if score >= 22:
                        for cell in row: cell.fill = purpleFill
                    if 8 <= score < 22:
                        for cell in row: cell.fill = greenFill
                    elif 6 <= score < 8:
                        for cell in row: cell.fill = lightGreenFill
                    elif 0 < score < 6:
                        for cell in row: cell.fill = yellowFill
                    elif score == 0:
                        for cell in row: cell.fill = whiteFill

print("✅ Script finished")